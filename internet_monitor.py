#!/usr/bin/env python3
"""
Internet connectivity monitor.

Runs continuously, checks connectivity every 15 minutes, and logs results
to Internet_Monitor.xlsx (one worksheet per day, retaining the latest 7 days).

The previous day's report is exported and uploaded to S3 at 00:30 AM each day.
Failed uploads are queued in pending_uploads/ and retried when connectivity
returns. All upload activity is recorded in the Upload_Audit worksheet.
"""

from __future__ import annotations

import logging
import os
import shutil
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import boto3
import requests
from botocore.exceptions import BotoCoreError, ClientError
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

WORKBOOK_NAME = "Internet_Monitor.xlsx"
CHECK_INTERVAL_MINUTES = 15
DAILY_UPLOAD_HOUR = 0
DAILY_UPLOAD_MINUTE = 30
RETENTION_DAYS = 7
CONNECTIVITY_URL = "https://www.google.com"
REQUEST_TIMEOUT_SECONDS = 10

CONNECTIVITY_HEADERS = ["Timestamp", "Connected"]
AUDIT_SHEET_NAME = "Upload_Audit"
AUDIT_HEADERS = ["Report Date", "Upload Attempt Time", "Status"]

DATE_FORMAT = "%Y-%m-%d"
TIMESTAMP_FORMAT = "%Y-%m-%d %H:%M:%S"

SCRIPT_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = SCRIPT_DIR / WORKBOOK_NAME
LOG_PATH = SCRIPT_DIR / "internet_monitor.log"
PENDING_UPLOADS_DIR = SCRIPT_DIR / "pending_uploads"

# AWS S3 settings (from environment variables)
S3_BUCKET = os.environ.get("S3_BUCKET", "")
S3_PREFIX = os.environ.get("S3_PREFIX", "internet-monitor/")
AWS_REGION = os.environ.get("AWS_REGION", "us-east-1")

# Tracks previous connectivity state for reconnect detection
_last_connectivity: Optional[bool] = None

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------


def setup_logging() -> None:
    """Configure file and console logging."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[
            logging.FileHandler(LOG_PATH, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )


# ---------------------------------------------------------------------------
# Connectivity check
# ---------------------------------------------------------------------------


def check_internet_connectivity() -> bool:
    """
    Determine whether internet is available by making an HTTP GET request.

    Returns True if the request succeeds with a 2xx status code, False otherwise.
    """
    try:
        response = requests.get(
            CONNECTIVITY_URL,
            timeout=REQUEST_TIMEOUT_SECONDS,
            headers={"User-Agent": "InternetMonitor/1.0"},
        )
        return response.ok
    except requests.RequestException as exc:
        logging.debug("Connectivity check failed: %s", exc)
        return False


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------


def _is_date_sheet_name(name: str) -> bool:
    """Return True if the sheet name matches YYYY-MM-DD."""
    try:
        datetime.strptime(name, DATE_FORMAT)
        return True
    except ValueError:
        return False


def _parse_sheet_date(name: str) -> date:
    """Parse a worksheet name into a date object."""
    return datetime.strptime(name, DATE_FORMAT).date()


def _get_date_sheets(workbook) -> list[tuple[date, str]]:
    """
    Return date-named worksheets sorted oldest to newest.

    Each entry is (sheet_date, sheet_name). Upload_Audit is excluded.
    """
    sheets: list[tuple[date, str]] = []
    for name in workbook.sheetnames:
        if _is_date_sheet_name(name):
            sheets.append((_parse_sheet_date(name), name))
    sheets.sort(key=lambda item: item[0])
    return sheets


def _ensure_workbook_exists() -> None:
    """
    Create the Excel workbook if missing, or ensure Upload_Audit exists.

    Upload_Audit is a permanent sheet with columns:
    Report Date | Upload Attempt Time | Status
    """
    if not WORKBOOK_PATH.exists():
        logging.info("Creating new workbook: %s", WORKBOOK_PATH)
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        today_name = date.today().strftime(DATE_FORMAT)
        sheet = workbook.create_sheet(title=today_name)
        _initialize_connectivity_sheet(sheet)
        _ensure_upload_audit_sheet(workbook)
        workbook.save(WORKBOOK_PATH)
        return

    workbook = load_workbook(WORKBOOK_PATH)
    try:
        had_audit = AUDIT_SHEET_NAME in workbook.sheetnames
        _ensure_upload_audit_sheet(workbook)
        if not had_audit:
            workbook.save(WORKBOOK_PATH)
    finally:
        workbook.close()


def _initialize_connectivity_sheet(sheet: Worksheet) -> None:
    """Write connectivity column headers on a new daily worksheet."""
    sheet.append(CONNECTIVITY_HEADERS)


def _ensure_upload_audit_sheet(workbook) -> Worksheet:
    """Return the Upload_Audit sheet, creating it with headers if missing."""
    if AUDIT_SHEET_NAME in workbook.sheetnames:
        return workbook[AUDIT_SHEET_NAME]

    sheet = workbook.create_sheet(title=AUDIT_SHEET_NAME)
    sheet.append(AUDIT_HEADERS)
    logging.info("Created %s worksheet", AUDIT_SHEET_NAME)
    return sheet


def _has_successful_upload(workbook, report_date: date) -> bool:
    """Return True if Upload_Audit already contains a SUCCESS row for report_date."""
    if AUDIT_SHEET_NAME not in workbook.sheetnames:
        return False

    report_str = report_date.strftime(DATE_FORMAT)
    sheet = workbook[AUDIT_SHEET_NAME]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == report_str and row[2] == "SUCCESS":
            return True
    return False


def append_audit_record(
    workbook,
    report_date: date,
    status: str,
    attempt_time: Optional[datetime] = None,
) -> None:
    """
    Append a new row to Upload_Audit (never overwrites existing records).

    status must be 'SUCCESS' or 'PENDING'.
    """
    sheet = _ensure_upload_audit_sheet(workbook)
    timestamp = (attempt_time or datetime.now()).strftime(TIMESTAMP_FORMAT)
    sheet.append([report_date.strftime(DATE_FORMAT), timestamp, status])


def _get_or_create_today_sheet(workbook) -> tuple[Worksheet, bool]:
    """
    Return today's worksheet and whether it was just created (new day rollover).

    Returns:
        (sheet, created_new_day)
    """
    today_name = date.today().strftime(DATE_FORMAT)

    if today_name in workbook.sheetnames:
        return workbook[today_name], False

    logging.info("Creating worksheet for new day: %s", today_name)
    sheet = workbook.create_sheet(title=today_name)
    _initialize_connectivity_sheet(sheet)
    return sheet, True


def _enforce_retention(workbook) -> None:
    """
    Keep at most RETENTION_DAYS date-named worksheets.

    When an eighth day would exceed the limit, remove the oldest sheet(s).
    Upload_Audit is never removed.
    """
    date_sheets = _get_date_sheets(workbook)

    while len(date_sheets) > RETENTION_DAYS:
        oldest_date, oldest_name = date_sheets.pop(0)
        logging.info(
            "Retention limit reached; removing oldest worksheet: %s (%s)",
            oldest_name,
            oldest_date,
        )
        del workbook[oldest_name]


def _append_result(sheet: Worksheet, connected: bool, check_time: datetime) -> None:
    """Append a connectivity result row using the scheduled boundary timestamp."""
    sheet.append([check_time.strftime(TIMESTAMP_FORMAT), connected])


def record_connectivity_result(connected: bool, check_time: datetime) -> None:
    """
    Open the workbook, append the latest check, enforce retention, and save.

    New day sheet creation is separate from the 00:30 scheduled upload.
    """
    _ensure_workbook_exists()

    workbook = load_workbook(WORKBOOK_PATH)
    try:
        _ensure_upload_audit_sheet(workbook)
        sheet, _created_new_day = _get_or_create_today_sheet(workbook)
        _append_result(sheet, connected, check_time)
        _enforce_retention(workbook)
        workbook.save(WORKBOOK_PATH)
    finally:
        workbook.close()


# ---------------------------------------------------------------------------
# Export and S3 upload
# ---------------------------------------------------------------------------


def export_day_sheet(workbook, report_date: date, output_path: Path) -> Path:
    """
    Export a single day's worksheet to a standalone Excel file.

    Copies all rows (headers + data) from the source sheet.
    """
    sheet_name = report_date.strftime(DATE_FORMAT)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet {sheet_name} not found in workbook")

    source = workbook[sheet_name]
    export_wb = Workbook()
    export_wb.remove(export_wb.active)
    export_sheet = export_wb.create_sheet(title=sheet_name)

    for row in source.iter_rows(values_only=True):
        export_sheet.append(list(row))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    export_wb.save(output_path)
    export_wb.close()
    logging.info("Exported report to %s", output_path)
    return output_path


def _get_s3_client():
    """Create a boto3 S3 client using environment variable credentials."""
    return boto3.client(
        "s3",
        region_name=AWS_REGION,
        aws_access_key_id=os.environ.get("AWS_ACCESS_KEY_ID"),
        aws_secret_access_key=os.environ.get("AWS_SECRET_ACCESS_KEY"),
    )


def upload_report_to_s3(local_path: Path, report_date: date) -> None:
    """Upload a report file to the configured S3 bucket."""
    if not S3_BUCKET:
        raise ValueError("S3_BUCKET environment variable is not set")

    s3_key = f"{S3_PREFIX}{report_date.strftime(DATE_FORMAT)}.xlsx"
    client = _get_s3_client()
    client.upload_file(str(local_path), S3_BUCKET, s3_key)
    logging.info(
        "Uploaded %s to s3://%s/%s",
        local_path.name,
        S3_BUCKET,
        s3_key,
    )


def _pending_file_path(report_date: date) -> Path:
    """Return the standard path for a pending upload file."""
    return PENDING_UPLOADS_DIR / f"{report_date.strftime(DATE_FORMAT)}.xlsx"


def _try_upload_and_audit(
    workbook,
    file_path: Path,
    report_date: date,
    *,
    is_pending_retry: bool = False,
) -> bool:
    """
    Attempt to upload a report file to S3 and record the result in Upload_Audit.

    On first-attempt failure: appends PENDING audit row and moves file to
    pending_uploads/. On pending retry failure: leaves file in place, no new
    audit row. On success: appends SUCCESS audit row.

    Returns True if upload succeeded.
    """
    attempt_time = datetime.now()

    try:
        upload_report_to_s3(file_path, report_date)
        append_audit_record(workbook, report_date, "SUCCESS", attempt_time)
        return True
    except (BotoCoreError, ClientError, ValueError, OSError):
        logging.exception("S3 upload failed for report %s", report_date)

        if not is_pending_retry:
            append_audit_record(workbook, report_date, "PENDING", attempt_time)
            dest = _pending_file_path(report_date)
            PENDING_UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
            if file_path.resolve() != dest.resolve():
                shutil.move(str(file_path), str(dest))
            logging.info("Report queued in %s", dest)

        return False


def _should_upload_yesterday_report(workbook, check_time: datetime) -> bool:
    """
    Return True when the previous day's report should be uploaded.

    Upload runs at 00:30 AM. If that slot was missed (e.g. app was down),
    upload on the next check the same day as long as no SUCCESS exists yet.
    """
    yesterday = check_time.date() - timedelta(days=1)
    sheet_name = yesterday.strftime(DATE_FORMAT)

    if sheet_name not in workbook.sheetnames:
        return False
    if _has_successful_upload(workbook, yesterday):
        return False

    upload_slot = check_time.replace(
        hour=DAILY_UPLOAD_HOUR,
        minute=DAILY_UPLOAD_MINUTE,
        second=0,
        microsecond=0,
    )
    return check_time >= upload_slot


def run_scheduled_previous_day_upload(check_time: datetime) -> None:
    """
    Export and upload the previous day's report at 00:30 AM (or if missed).

    If upload fails, the file is stored in pending_uploads/ and monitoring
    continues without interruption.
    """
    _ensure_workbook_exists()

    workbook = load_workbook(WORKBOOK_PATH)
    try:
        if not _should_upload_yesterday_report(workbook, check_time):
            return

        yesterday = check_time.date() - timedelta(days=1)
        logging.info(
            "Scheduled upload at %02d:%02d for report %s",
            DAILY_UPLOAD_HOUR,
            DAILY_UPLOAD_MINUTE,
            yesterday.strftime(DATE_FORMAT),
        )
        handle_day_rollover_upload(workbook, yesterday)
        workbook.save(WORKBOOK_PATH)
    except Exception:
        logging.exception("Scheduled previous-day upload failed; monitoring continues")
    finally:
        workbook.close()


def handle_day_rollover_upload(workbook, report_date: date) -> None:
    """
    Finalize, export, and upload a daily report to S3.

    If upload fails, the file is stored in pending_uploads/ and monitoring
    continues without interruption.
    """
    sheet_name = report_date.strftime(DATE_FORMAT)
    if sheet_name not in workbook.sheetnames:
        logging.info("No sheet for %s; skipping day-rollover upload", sheet_name)
        return

    logging.info("Finalizing report for %s", sheet_name)

    try:
        temp_path = PENDING_UPLOADS_DIR / f".tmp_{sheet_name}.xlsx"
        export_day_sheet(workbook, report_date, temp_path)

        if _try_upload_and_audit(workbook, temp_path, report_date):
            if temp_path.exists():
                temp_path.unlink()
    except Exception:
        logging.exception(
            "Day rollover export/upload failed for %s; monitoring continues",
            sheet_name,
        )


def process_pending_uploads(workbook) -> None:
    """
    Scan pending_uploads/ and upload all queued report files to S3.

    Successfully uploaded files are removed from pending_uploads/.
    Each success appends a new SUCCESS row to Upload_Audit.
    """
    if not PENDING_UPLOADS_DIR.exists():
        return

    pending_files = sorted(
        p for p in PENDING_UPLOADS_DIR.glob("*.xlsx") if not p.name.startswith(".tmp_")
    )
    if not pending_files:
        return

    logging.info("Processing %d pending upload(s)", len(pending_files))

    for file_path in pending_files:
        try:
            report_date = datetime.strptime(file_path.stem, DATE_FORMAT).date()
        except ValueError:
            logging.warning("Skipping unrecognized pending file: %s", file_path.name)
            continue

        try:
            if _try_upload_and_audit(
                workbook, file_path, report_date, is_pending_retry=True
            ):
                file_path.unlink(missing_ok=True)
                logging.info("Removed uploaded file from pending_uploads: %s", file_path.name)
        except Exception:
            logging.exception(
                "Error processing pending upload %s; continuing with next file",
                file_path.name,
            )


def retry_pending_uploads() -> None:
    """Open the workbook, process pending uploads, and save audit records."""
    if not PENDING_UPLOADS_DIR.exists():
        return

    has_pending = any(
        not p.name.startswith(".tmp_") for p in PENDING_UPLOADS_DIR.glob("*.xlsx")
    )
    if not has_pending:
        return

    try:
        workbook = load_workbook(WORKBOOK_PATH)
        try:
            process_pending_uploads(workbook)
            workbook.save(WORKBOOK_PATH)
        finally:
            workbook.close()
    except Exception:
        logging.exception("Failed to process pending uploads")


# ---------------------------------------------------------------------------
# Scheduled job
# ---------------------------------------------------------------------------


def run_connectivity_check(check_time: datetime) -> None:
    """Perform one connectivity check, persist result, and retry pending uploads."""
    global _last_connectivity

    try:
        connected = check_internet_connectivity()
        status = "CONNECTED" if connected else "DISCONNECTED"
        logging.info(
            "Internet check at %s: %s",
            check_time.strftime(TIMESTAMP_FORMAT),
            status,
        )

        record_connectivity_result(connected, check_time)
        logging.info("Result saved to %s", WORKBOOK_PATH)

        run_scheduled_previous_day_upload(check_time)

        if connected:
            if _last_connectivity is False:
                logging.info(
                    "Internet connectivity restored; scanning pending_uploads"
                )
            retry_pending_uploads()

        _last_connectivity = connected
    except Exception:
        logging.exception("Connectivity check or file operation failed; continuing.")


def _current_boundary(now: datetime) -> datetime:
    """Return the 15-minute boundary at or before `now` (seconds zeroed)."""
    minute_slot = (now.minute // CHECK_INTERVAL_MINUTES) * CHECK_INTERVAL_MINUTES
    return now.replace(minute=minute_slot, second=0, microsecond=0)


def _initial_check_time() -> datetime:
    """
    Return the first scheduled check time on startup.

    If the process starts within 1 second of a boundary, use that boundary;
    otherwise use the next upcoming boundary.
    """
    now = datetime.now()
    boundary = _current_boundary(now)
    elapsed = (now - boundary).total_seconds()
    if elapsed < 1.0:
        return boundary
    return boundary + timedelta(minutes=CHECK_INTERVAL_MINUTES)


def _sleep_until(target: datetime) -> None:
    """Block until `target`, logging the wait duration."""
    sleep_for = (target - datetime.now()).total_seconds()
    if sleep_for <= 0:
        return
    logging.info(
        "Next check at %s; sleeping %.0f seconds",
        target.strftime(TIMESTAMP_FORMAT),
        sleep_for,
    )
    time.sleep(sleep_for)


# ---------------------------------------------------------------------------
# Main loop
# ---------------------------------------------------------------------------


def main() -> None:
    """Entry point: set up logging, run checks on exact 15-minute boundaries."""
    setup_logging()
    logging.info("Internet monitor starting")
    logging.info(
        "Workbook: %s | Interval: %d min | Upload: %02d:%02d | Retention: %d days | S3: %s",
        WORKBOOK_PATH,
        CHECK_INTERVAL_MINUTES,
        DAILY_UPLOAD_HOUR,
        DAILY_UPLOAD_MINUTE,
        RETENTION_DAYS,
        S3_BUCKET or "(not configured)",
    )

    if not S3_BUCKET:
        logging.warning(
            "S3_BUCKET is not set; uploads will fail and reports will be queued "
            "in pending_uploads/"
        )

    # Catch up if the app started after today's 00:30 upload slot.
    run_scheduled_previous_day_upload(datetime.now())

    next_check = _initial_check_time()

    while True:
        try:
            _sleep_until(next_check)
            run_connectivity_check(next_check)
            next_check += timedelta(minutes=CHECK_INTERVAL_MINUTES)
        except KeyboardInterrupt:
            logging.info("Shutdown requested; exiting.")
            break
        except Exception:
            logging.exception("Unexpected error in main loop; continuing.")
            next_check += timedelta(minutes=CHECK_INTERVAL_MINUTES)
            time.sleep(5)


if __name__ == "__main__":
    main()
